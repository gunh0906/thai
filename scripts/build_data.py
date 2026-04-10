from __future__ import annotations

import argparse
import json
import re
from datetime import datetime
from pathlib import Path
from typing import Iterable

import openpyxl
from expanded_more import MORE_SENTENCES, MORE_VOCAB
from expanded_terms import EXPANDED_SENTENCES, EXPANDED_VOCAB
from generated_bulk import generate_bulk_entries

SCENARIOS = [
    {
        "id": "all",
        "label": "전체",
        "description": "모든 단어와 문장을 함께 봅니다.",
    },
    {
        "id": "기본회화",
        "label": "기본회화",
        "description": "대답, 부탁, 이해 여부처럼 자주 쓰는 표현",
    },
    {
        "id": "인사",
        "label": "인사",
        "description": "자기소개, 감사, 사과, 작별 인사",
    },
    {
        "id": "식당",
        "label": "식당",
        "description": "주문, 계산, 맛 표현, 포장",
    },
    {
        "id": "이동",
        "label": "이동",
        "description": "길 찾기, 방향, 위치, 이동 관련 표현",
    },
    {
        "id": "쇼핑",
        "label": "쇼핑",
        "description": "가격, 색상, 사이즈, 결제 관련 표현",
    },
    {
        "id": "건강",
        "label": "건강",
        "description": "몸 상태, 약, 병원, 도움 요청",
    },
    {
        "id": "일터",
        "label": "일터",
        "description": "확인, 대기, 완료, 속도, 문제 상황",
    },
    {
        "id": "숫자·시간",
        "label": "숫자·시간",
        "description": "시간, 요일, 날짜 흐름 관련 표현",
    },
]

SCENARIO_ORDER = {item["id"]: index for index, item in enumerate(SCENARIOS)}

TAG_RULES = {
    "기본회화": [
        "어떻게",
        "뭐",
        "이거",
        "저는",
        "입니다",
        "가능",
        "불가능",
        "사용",
        "이해",
        "알겠습니다",
        "몰라",
        "다시",
        "천천히",
        "도와",
        "잠깐",
        "괜찮",
    ],
    "인사": [
        "안녕",
        "반가워",
        "잘가",
        "계세요",
        "감사",
        "죄송",
        "실례",
        "이름",
        "사람",
        "한국",
        "태국",
        "소개",
    ],
    "식당": [
        "먹",
        "밥",
        "메뉴",
        "계산",
        "포장",
        "물",
        "맛",
        "맵",
        "음식",
        "주문",
        "식당",
        "봉투",
    ],
    "이동": [
        "가다",
        "오다",
        "올라가다",
        "내려가다",
        "어디",
        "왼쪽",
        "오른쪽",
        "위치",
        "길",
        "집",
        "근처",
        "가깝",
        "멀",
    ],
    "쇼핑": [
        "얼마",
        "비싸",
        "깎",
        "카드",
        "현금",
        "사이즈",
        "색",
        "사다",
        "결제",
        "가격",
    ],
    "건강": [
        "아프",
        "기침",
        "감기",
        "두통",
        "병원",
        "약",
        "위험",
        "조심",
        "도와",
        "화장실",
        "열",
        "응급",
    ],
    "일터": [
        "중요",
        "먼저",
        "가져오다",
        "측정",
        "결과",
        "기억",
        "집중",
        "문제",
        "확인",
        "완료",
        "끝",
        "기다리",
        "빠르",
        "느리",
        "다시 해",
        "작업",
    ],
    "숫자·시간": [
        "시",
        "점심",
        "매일",
        "오늘",
        "내일",
        "어제",
        "주",
        "월요일",
        "화요일",
        "수요일",
        "목요일",
        "금요일",
        "토요일",
        "일요일",
        "십",
        "백",
        "천",
        "만",
    ],
}

STOPWORDS = {
    "",
    "입니다",
    "있다",
    "하다",
    "할",
    "수",
    "있어요",
    "있습니다",
    "주세요",
    "조금",
    "너무",
    "정말",
    "저는",
    "그는",
    "이거",
    "저거",
    "이걸로",
    "좀",
    "및",
    "그리고",
    "에서",
    "으로",
    "에게",
    "이다",
    "안",
    "못",
    "해요",
    "캅",
    "카",
}

SUPPLEMENTAL_VOCAB = [
    {
        "thai": "사왓디 캅/카",
        "korean": "안녕하세요",
        "tags": ["인사", "기본회화"],
        "note": "남성은 캅, 여성은 카를 붙이면 더 자연스럽습니다.",
    },
    {
        "thai": "컵쿤 캅/카",
        "korean": "감사합니다",
        "tags": ["인사", "기본회화"],
    },
    {
        "thai": "커톳 캅/카",
        "korean": "실례합니다 / 죄송합니다",
        "tags": ["인사", "기본회화"],
    },
    {
        "thai": "마이 벤 라이",
        "korean": "괜찮아요 / 천만에요",
        "tags": ["기본회화"],
    },
    {
        "thai": "추어 아라이",
        "korean": "이름이 뭐예요?",
        "tags": ["인사", "기본회화"],
    },
    {
        "thai": "폼 추어 ...",
        "korean": "제 이름은 ... 입니다",
        "tags": ["인사", "기본회화"],
    },
    {
        "thai": "푸트 파사 앙끄릿 다이 마이",
        "korean": "영어 할 수 있어요?",
        "tags": ["기본회화"],
    },
    {
        "thai": "푸트 차차 너이",
        "korean": "천천히 말해주세요",
        "tags": ["기본회화"],
    },
    {
        "thai": "푸트 이끈 티 너이",
        "korean": "한 번 더 말해주세요",
        "tags": ["기본회화"],
    },
    {
        "thai": "츄어이 두어이",
        "korean": "도와주세요",
        "tags": ["기본회화", "건강"],
    },
    {
        "thai": "러 사끄루",
        "korean": "잠깐만요",
        "tags": ["기본회화"],
    },
    {
        "thai": "다이",
        "korean": "할 수 있다 / 가능하다",
        "tags": ["기본회화"],
    },
    {
        "thai": "마이 다이",
        "korean": "안 된다 / 못 한다",
        "tags": ["기본회화"],
    },
    {
        "thai": "아로이",
        "korean": "맛있다",
        "tags": ["식당"],
    },
    {
        "thai": "펫",
        "korean": "맵다",
        "tags": ["식당", "건강"],
    },
    {
        "thai": "마이 펫",
        "korean": "안 맵게",
        "tags": ["식당"],
    },
    {
        "thai": "남 쁠라오",
        "korean": "생수 / 물",
        "tags": ["식당"],
    },
    {
        "thai": "남 캥",
        "korean": "얼음",
        "tags": ["식당"],
    },
    {
        "thai": "아한",
        "korean": "음식",
        "tags": ["식당"],
    },
    {
        "thai": "긴",
        "korean": "먹다",
        "tags": ["식당"],
    },
    {
        "thai": "아오",
        "korean": "원하다 / 주세요",
        "tags": ["식당", "쇼핑"],
    },
    {
        "thai": "아오 안니",
        "korean": "이걸로 주세요",
        "tags": ["식당", "쇼핑"],
    },
    {
        "thai": "싸이 퉁",
        "korean": "봉투에 넣다 / 포장하다",
        "tags": ["식당", "쇼핑"],
    },
    {
        "thai": "체크 빈",
        "korean": "계산서 / 계산",
        "tags": ["식당"],
    },
    {
        "thai": "아로이 막",
        "korean": "정말 맛있다",
        "tags": ["식당"],
    },
    {
        "thai": "유 티 나이",
        "korean": "어디에 있어요?",
        "tags": ["이동", "기본회화"],
    },
    {
        "thai": "트롱 빠이",
        "korean": "쭉 가세요",
        "tags": ["이동"],
    },
    {
        "thai": "리여우 싸이",
        "korean": "왼쪽으로 도세요",
        "tags": ["이동"],
    },
    {
        "thai": "리여우 크와",
        "korean": "오른쪽으로 도세요",
        "tags": ["이동"],
    },
    {
        "thai": "끌라이",
        "korean": "가깝다",
        "tags": ["이동"],
    },
    {
        "thai": "글라이",
        "korean": "멀다",
        "tags": ["이동"],
    },
    {
        "thai": "톤니",
        "korean": "지금",
        "tags": ["이동", "숫자·시간"],
    },
    {
        "thai": "프룽니",
        "korean": "내일",
        "tags": ["숫자·시간"],
    },
    {
        "thai": "므어완니",
        "korean": "어제",
        "tags": ["숫자·시간"],
    },
    {
        "thai": "차오",
        "korean": "아침",
        "tags": ["숫자·시간"],
    },
    {
        "thai": "옌",
        "korean": "저녁",
        "tags": ["숫자·시간"],
    },
    {
        "thai": "타오라이",
        "korean": "얼마예요?",
        "tags": ["쇼핑"],
    },
    {
        "thai": "팽 빠이",
        "korean": "너무 비싸요",
        "tags": ["쇼핑"],
    },
    {
        "thai": "롯 너이 다이 마이",
        "korean": "조금 깎아줄 수 있어요?",
        "tags": ["쇼핑"],
    },
    {
        "thai": "씨",
        "korean": "색",
        "tags": ["쇼핑"],
    },
    {
        "thai": "사이",
        "korean": "사이즈",
        "tags": ["쇼핑"],
    },
    {
        "thai": "카드 다이 마이",
        "korean": "카드 돼요?",
        "tags": ["쇼핑"],
    },
    {
        "thai": "미 씨 은 마이",
        "korean": "다른 색 있어요?",
        "tags": ["쇼핑"],
    },
    {
        "thai": "렉 껀 빠이",
        "korean": "너무 작아요",
        "tags": ["쇼핑"],
    },
    {
        "thai": "야이 껀 빠이",
        "korean": "너무 커요",
        "tags": ["쇼핑"],
    },
    {
        "thai": "쨉",
        "korean": "아프다",
        "tags": ["건강"],
    },
    {
        "thai": "푸앗 후아",
        "korean": "머리가 아프다",
        "tags": ["건강"],
    },
    {
        "thai": "푸앗 통",
        "korean": "배가 아프다",
        "tags": ["건강"],
    },
    {
        "thai": "벤 카이",
        "korean": "열이 나다",
        "tags": ["건강"],
    },
    {
        "thai": "롱 파야반",
        "korean": "병원",
        "tags": ["건강"],
    },
    {
        "thai": "야",
        "korean": "약",
        "tags": ["건강"],
    },
    {
        "thai": "홍 남",
        "korean": "화장실",
        "tags": ["건강", "이동"],
    },
    {
        "thai": "로 껀",
        "korean": "잠깐 기다리세요",
        "tags": ["일터", "기본회화"],
    },
    {
        "thai": "름 다이 러이",
        "korean": "바로 시작해도 됩니다",
        "tags": ["일터"],
    },
    {
        "thai": "윳 껀",
        "korean": "먼저 멈추세요",
        "tags": ["일터", "건강"],
    },
    {
        "thai": "트루앗 이끈 크랑",
        "korean": "한 번 더 확인하세요",
        "tags": ["일터"],
    },
    {
        "thai": "미 빤하",
        "korean": "문제가 있다",
        "tags": ["일터"],
    },
    {
        "thai": "리압러이",
        "korean": "완료됐다 / 정리됐다",
        "tags": ["일터"],
    },
    {
        "thai": "양 마이 셋",
        "korean": "아직 안 끝났다",
        "tags": ["일터"],
    },
    {
        "thai": "레오 빠이",
        "korean": "너무 빠르다",
        "tags": ["일터"],
    },
    {
        "thai": "차 빠이",
        "korean": "너무 느리다",
        "tags": ["일터"],
    },
    {
        "thai": "롱 이끈 크랑",
        "korean": "다시 한 번 해보다",
        "tags": ["일터"],
    },
    {
        "thai": "찬 니",
        "korean": "이 시간 / 이때",
        "tags": ["숫자·시간"],
    },
]

SUPPLEMENTAL_SENTENCES = [
    {
        "thai": "사왓디 캅 폼 마짝 까올리",
        "korean": "안녕하세요, 저는 한국에서 왔어요.",
        "tags": ["인사", "기본회화"],
    },
    {
        "thai": "컵쿤 막 캅",
        "korean": "정말 감사합니다.",
        "tags": ["인사", "기본회화"],
    },
    {
        "thai": "커톳 캅 푸트 차차 너이",
        "korean": "죄송하지만 천천히 말해주세요.",
        "tags": ["기본회화"],
    },
    {
        "thai": "폼 마이 카우짜이 푸트 이끈 티 너이",
        "korean": "제가 이해를 못했어요. 한 번 더 말해주세요.",
        "tags": ["기본회화"],
    },
    {
        "thai": "츄어이 두어이 나 캅",
        "korean": "좀 도와주세요.",
        "tags": ["기본회화", "건강"],
    },
    {
        "thai": "러 사끄루 나 캅",
        "korean": "잠깐만 기다려 주세요.",
        "tags": ["기본회화"],
    },
    {
        "thai": "추어 아라이 캅",
        "korean": "이름이 뭐예요?",
        "tags": ["인사", "기본회화"],
    },
    {
        "thai": "폼 추어 민수 캅",
        "korean": "제 이름은 민수예요.",
        "tags": ["인사", "기본회화"],
    },
    {
        "thai": "마이 벤 라이 캅",
        "korean": "괜찮아요.",
        "tags": ["기본회화"],
    },
    {
        "thai": "푸트 파사 앙끄릿 다이 마이 캅",
        "korean": "영어 할 수 있으세요?",
        "tags": ["기본회화"],
    },
    {
        "thai": "커 남 쁠라오 능 쿠앗 캅",
        "korean": "물 한 병 주세요.",
        "tags": ["식당"],
    },
    {
        "thai": "마이 펫 너이 나 캅",
        "korean": "안 맵게 부탁해요.",
        "tags": ["식당"],
    },
    {
        "thai": "아한 찬니 아로이 막",
        "korean": "이 음식 정말 맛있어요.",
        "tags": ["식당"],
    },
    {
        "thai": "체크 빈 너이 캅",
        "korean": "계산 부탁해요.",
        "tags": ["식당"],
    },
    {
        "thai": "싸이 퉁 하이 너이 캅",
        "korean": "봉투에 넣어 주세요.",
        "tags": ["식당", "쇼핑"],
    },
    {
        "thai": "커 남 캥 너이 캅",
        "korean": "얼음 조금 주세요.",
        "tags": ["식당"],
    },
    {
        "thai": "아오 안니 캅",
        "korean": "이걸로 주세요.",
        "tags": ["식당", "쇼핑"],
    },
    {
        "thai": "낀 티니 캅",
        "korean": "여기서 먹을게요.",
        "tags": ["식당"],
    },
    {
        "thai": "아오 글랍 반 캅",
        "korean": "포장해 주세요.",
        "tags": ["식당"],
    },
    {
        "thai": "유 티 나이 캅",
        "korean": "어디에 있어요?",
        "tags": ["이동", "기본회화"],
    },
    {
        "thai": "트롱 빠이 레오 리여우 싸이",
        "korean": "쭉 가다가 왼쪽으로 도세요.",
        "tags": ["이동"],
    },
    {
        "thai": "트롱 빠이 레오 리여우 크와",
        "korean": "쭉 가다가 오른쪽으로 도세요.",
        "tags": ["이동"],
    },
    {
        "thai": "티니 끌라이 마이 캅",
        "korean": "여기 가까워요?",
        "tags": ["이동"],
    },
    {
        "thai": "톤니 폼 짜 깝 반",
        "korean": "지금 저는 집에 갈 거예요.",
        "tags": ["이동", "숫자·시간"],
    },
    {
        "thai": "프룽니 쪼 깐 다이 마이",
        "korean": "내일 만날 수 있을까요?",
        "tags": ["숫자·시간", "인사"],
    },
    {
        "thai": "홍 남 유 티 나이 캅",
        "korean": "화장실이 어디예요?",
        "tags": ["건강", "이동"],
    },
    {
        "thai": "안니 타오라이 캅",
        "korean": "이거 얼마예요?",
        "tags": ["쇼핑"],
    },
    {
        "thai": "팽 빠이 롯 너이 다이 마이",
        "korean": "너무 비싼데 조금 깎아주실 수 있어요?",
        "tags": ["쇼핑"],
    },
    {
        "thai": "미 씨 은 마이 캅",
        "korean": "다른 색 있어요?",
        "tags": ["쇼핑"],
    },
    {
        "thai": "야이 꽌 니 미 마이",
        "korean": "이보다 큰 사이즈 있나요?",
        "tags": ["쇼핑"],
    },
    {
        "thai": "렉 꽌 니 미 마이",
        "korean": "이보다 작은 사이즈 있나요?",
        "tags": ["쇼핑"],
    },
    {
        "thai": "카드 다이 마이 캅",
        "korean": "카드 결제 되나요?",
        "tags": ["쇼핑"],
    },
    {
        "thai": "폼 쨉 후아 너이",
        "korean": "저 머리가 좀 아파요.",
        "tags": ["건강"],
    },
    {
        "thai": "폼 푸앗 통",
        "korean": "배가 아파요.",
        "tags": ["건강"],
    },
    {
        "thai": "폼 벤 카이",
        "korean": "열이 나는 것 같아요.",
        "tags": ["건강"],
    },
    {
        "thai": "롱 파야반 유 티 나이",
        "korean": "병원이 어디에 있어요?",
        "tags": ["건강", "이동"],
    },
    {
        "thai": "커 야 다이 마이",
        "korean": "약을 받을 수 있을까요?",
        "tags": ["건강"],
    },
    {
        "thai": "라우앙 너이 나",
        "korean": "조심해 주세요.",
        "tags": ["건강", "일터"],
    },
    {
        "thai": "로 껀 나 캅",
        "korean": "잠깐만 기다려 주세요.",
        "tags": ["일터", "기본회화"],
    },
    {
        "thai": "름 다이 러이",
        "korean": "바로 시작해도 됩니다.",
        "tags": ["일터"],
    },
    {
        "thai": "윳 껀 곤",
        "korean": "먼저 멈춰 주세요.",
        "tags": ["일터", "건강"],
    },
    {
        "thai": "미 빤하 티니",
        "korean": "여기에 문제가 있어요.",
        "tags": ["일터"],
    },
    {
        "thai": "트루앗 이끈 크랑 나 캅",
        "korean": "한 번 더 확인해 주세요.",
        "tags": ["일터"],
    },
    {
        "thai": "양 마이 셋",
        "korean": "아직 안 끝났어요.",
        "tags": ["일터"],
    },
    {
        "thai": "리압러이 레오 캅",
        "korean": "이제 완료됐습니다.",
        "tags": ["일터"],
    },
    {
        "thai": "레오 빠이 너이",
        "korean": "조금 너무 빨라요.",
        "tags": ["일터"],
    },
    {
        "thai": "차 빠이 너이",
        "korean": "조금 너무 느려요.",
        "tags": ["일터"],
    },
    {
        "thai": "롱 이끈 크랑 다이 마이",
        "korean": "다시 한 번 해볼 수 있을까요?",
        "tags": ["일터"],
    },
    {
        "thai": "완니 융 막",
        "korean": "오늘 정말 바빠요.",
        "tags": ["일터", "숫자·시간"],
    },
    {
        "thai": "프룽니 차오 쪼 깐",
        "korean": "내일 아침에 만나요.",
        "tags": ["숫자·시간", "인사"],
    },
]


def clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = re.sub(r"\s+", " ", text)
    return text


def normalize_for_id(value: str) -> str:
    text = clean_text(value).lower()
    text = re.sub(r"[^0-9a-z가-힣]+", "-", text)
    text = re.sub(r"-{2,}", "-", text).strip("-")
    return text or "item"


def normalize_compact(value: str) -> str:
    text = clean_text(value).lower()
    text = re.sub(r"[^0-9a-z가-힣\u0E00-\u0E7F]+", "", text)
    return text


def split_korean_variants(value: str) -> list[str]:
    parts = [clean_text(part) for part in re.split(r"[/|,·]", clean_text(value)) if clean_text(part)]
    if not parts:
        return []
    variants: list[str] = []
    for part in parts:
        compact = normalize_compact(part)
        if compact and compact not in variants:
            variants.append(compact)
    return variants


def extract_keywords(*parts: str) -> list[str]:
    text = " ".join(clean_text(part) for part in parts if clean_text(part))
    tokens: list[str] = []
    for token in re.split(r"[\s,./()\-?!=:+]+", text.lower()):
        token = token.strip()
        if len(token) < 2 or token in STOPWORDS:
            continue
        if token not in tokens:
            tokens.append(token)
    return tokens


def detect_tags(*parts: str) -> list[str]:
    text = " ".join(clean_text(part) for part in parts if clean_text(part))
    tags: list[str] = []
    for tag, keywords in TAG_RULES.items():
        if any(keyword in text for keyword in keywords):
            tags.append(tag)
    if not tags:
        tags.append("기본회화")
    unique_tags = sorted(set(tags), key=lambda item: SCENARIO_ORDER.get(item, 999))
    return unique_tags


def make_entry(
    *,
    kind: str,
    source: str,
    sheet: str,
    index: int,
    thai: str,
    thai_script: str = "",
    korean: str,
    tags: Iterable[str] | None = None,
    note: str = "",
    extra_keywords: Iterable[str] | None = None,
) -> dict:
    clean_thai = clean_text(thai)
    clean_thai_script = clean_text(thai_script)
    clean_korean = clean_text(korean)
    clean_note = clean_text(note)
    derived_tags = list(tags) if tags else detect_tags(clean_thai, clean_korean, clean_note)
    keywords = extract_keywords(clean_thai, clean_thai_script, clean_korean, clean_note, " ".join(derived_tags))
    if extra_keywords:
        for item in extra_keywords:
            keyword = clean_text(item).lower()
            if len(keyword) < 2 or keyword in STOPWORDS or keyword in keywords:
                continue
            keywords.append(keyword)
    return {
        "id": f"{source}-{kind}-{index:03d}-{normalize_for_id(clean_thai or clean_korean)}",
        "kind": kind,
        "source": source,
        "sourceLabel": {
            "excel": "엑셀",
            "supplemental": "확장",
            "custom": "내가 추가",
        }.get(source, source),
        "sheet": sheet,
        "thai": clean_thai,
        "thaiScript": clean_thai_script,
        "korean": clean_korean,
        "note": clean_note,
        "tags": derived_tags,
        "keywords": keywords,
    }


def parse_vocab_sheet(ws) -> list[dict]:
    entries: list[dict] = []
    counter = 1
    for row in ws.iter_rows(min_row=2, values_only=True):
        for start in (0, 3):
            thai = clean_text(row[start + 1] if len(row) > start + 1 else "")
            korean = clean_text(row[start + 2] if len(row) > start + 2 else "")
            if not thai and not korean:
                continue
            entries.append(
                make_entry(
                    kind="vocab",
                    source="excel",
                    sheet=ws.title,
                    index=counter,
                    thai=thai,
                    korean=korean,
                )
            )
            counter += 1
    return entries


def parse_sentence_sheet(ws) -> list[dict]:
    entries: list[dict] = []
    counter = 1
    for row in ws.iter_rows(min_row=2, values_only=True):
        thai = clean_text(row[1] if len(row) > 1 else "")
        korean = clean_text(row[2] if len(row) > 2 else "")
        if not thai and not korean:
            continue
        entries.append(
            make_entry(
                kind="sentence",
                source="excel",
                sheet=ws.title,
                index=counter,
                thai=thai,
                korean=korean,
            )
        )
        counter += 1
    return entries


def parse_workbook(path: Path) -> tuple[list[dict], list[dict]]:
    workbook = openpyxl.load_workbook(path, data_only=True)
    vocab_entries: list[dict] = []
    sentence_entries: list[dict] = []
    for ws in workbook.worksheets:
        title = clean_text(ws.title)
        if title.startswith("단어"):
            vocab_entries.extend(parse_vocab_sheet(ws))
        elif title.startswith("문장"):
            sentence_entries.extend(parse_sentence_sheet(ws))
    return vocab_entries, sentence_entries


def build_supplemental_entries() -> tuple[list[dict], list[dict]]:
    vocab_entries = [
        make_entry(
            kind="vocab",
            source="supplemental",
            sheet="확장 단어",
            index=index,
            thai=item["thai"],
            thai_script=item.get("thaiScript", ""),
            korean=item["korean"],
            tags=item.get("tags"),
            note=item.get("note", ""),
            extra_keywords=item.get("keywords"),
        )
        for index, item in enumerate(SUPPLEMENTAL_VOCAB, start=1)
    ]
    vocab_entries.extend(
        [
            make_entry(
                kind="vocab",
                source="supplemental",
                sheet="확장 단어",
                index=len(vocab_entries) + index,
                thai=item["thai"],
                thai_script=item.get("thaiScript", ""),
                korean=item["korean"],
                tags=item.get("tags"),
                note=item.get("note", ""),
                extra_keywords=item.get("keywords"),
            )
            for index, item in enumerate(EXPANDED_VOCAB, start=1)
        ]
    )
    vocab_entries.extend(
        [
            make_entry(
                kind="vocab",
                source="supplemental",
                sheet="확장 단어",
                index=len(vocab_entries) + index,
                thai=item["thai"],
                thai_script=item.get("thaiScript", ""),
                korean=item["korean"],
                tags=item.get("tags"),
                note=item.get("note", ""),
                extra_keywords=item.get("keywords"),
            )
            for index, item in enumerate(MORE_VOCAB, start=1)
        ]
    )
    sentence_entries = [
        make_entry(
            kind="sentence",
            source="supplemental",
            sheet="확장 문장",
            index=index,
            thai=item["thai"],
            thai_script=item.get("thaiScript", ""),
            korean=item["korean"],
            tags=item.get("tags"),
            note=item.get("note", ""),
            extra_keywords=item.get("keywords"),
        )
        for index, item in enumerate(SUPPLEMENTAL_SENTENCES, start=1)
    ]
    sentence_entries.extend(
        [
            make_entry(
                kind="sentence",
                source="supplemental",
                sheet="확장 문장",
                index=len(sentence_entries) + index,
                thai=item["thai"],
                thai_script=item.get("thaiScript", ""),
                korean=item["korean"],
                tags=item.get("tags"),
                note=item.get("note", ""),
                extra_keywords=item.get("keywords"),
            )
            for index, item in enumerate(EXPANDED_SENTENCES, start=1)
        ]
    )
    sentence_entries.extend(
        [
            make_entry(
                kind="sentence",
                source="supplemental",
                sheet="확장 문장",
                index=len(sentence_entries) + index,
                thai=item["thai"],
                thai_script=item.get("thaiScript", ""),
                korean=item["korean"],
                tags=item.get("tags"),
                note=item.get("note", ""),
                extra_keywords=item.get("keywords"),
            )
            for index, item in enumerate(MORE_SENTENCES, start=1)
        ]
    )
    return vocab_entries, sentence_entries


def entry_preference_score(entry: dict) -> tuple[int, int, int, int]:
    has_script = int(bool(clean_text(entry.get("thaiScript", ""))))
    has_note = int(bool(clean_text(entry.get("note", ""))))
    keyword_count = len(entry.get("keywords", []))
    korean_penalty = -len(clean_text(entry.get("korean", "")))
    return (has_script, has_note, keyword_count, korean_penalty)


def entries_overlap(left: dict, right: dict) -> bool:
    left_thai = normalize_compact(left.get("thaiScript") or left.get("thai", ""))
    right_thai = normalize_compact(right.get("thaiScript") or right.get("thai", ""))
    if left_thai and right_thai and left_thai == right_thai:
        return True

    left_korean = normalize_compact(left.get("korean", ""))
    right_korean = normalize_compact(right.get("korean", ""))
    if left_korean and right_korean and left_korean == right_korean:
        return True

    left_variants = split_korean_variants(left.get("korean", ""))
    right_variants = split_korean_variants(right.get("korean", ""))
    if left_thai and right_thai and left_variants and right_variants:
        if any(item in right_variants for item in left_variants):
            return True

    return False


def merge_entries(base: dict, incoming: dict) -> dict:
    preferred = base if entry_preference_score(base) >= entry_preference_score(incoming) else incoming
    other = incoming if preferred is base else base

    merged = dict(preferred)
    if not clean_text(merged.get("thaiScript", "")):
        merged["thaiScript"] = clean_text(other.get("thaiScript", ""))
    if not clean_text(merged.get("note", "")):
        merged["note"] = clean_text(other.get("note", ""))

    merged["tags"] = sorted(
        set((preferred.get("tags") or []) + (other.get("tags") or [])),
        key=lambda item: SCENARIO_ORDER.get(item, 999),
    )
    merged["keywords"] = list(dict.fromkeys((preferred.get("keywords") or []) + (other.get("keywords") or [])))
    return merged


def deduplicate_entries(entries: list[dict]) -> list[dict]:
    deduped: list[dict] = []
    thai_index: dict[tuple[str, str], int] = {}
    korean_index: dict[tuple[str, str], int] = {}

    def register(index: int, entry: dict) -> None:
        thai_key = normalize_compact(entry.get("thaiScript") or entry.get("thai", ""))
        korean_key = normalize_compact(entry.get("korean", ""))
        kind_key = entry["kind"]
        if thai_key:
            thai_index[(kind_key, thai_key)] = index
        if korean_key:
            korean_index[(kind_key, korean_key)] = index

    for entry in entries:
        kind_key = entry["kind"]
        thai_key = normalize_compact(entry.get("thaiScript") or entry.get("thai", ""))
        korean_key = normalize_compact(entry.get("korean", ""))
        match_index = None

        if thai_key:
            match_index = thai_index.get((kind_key, thai_key))
        if match_index is None and korean_key:
            match_index = korean_index.get((kind_key, korean_key))

        if match_index is None:
            deduped.append(entry)
            register(len(deduped) - 1, entry)
            continue

        deduped[match_index] = merge_entries(deduped[match_index], entry)
        register(match_index, deduped[match_index])

    return deduped


def build_data(workbook_path: Path) -> dict:
    excel_vocab, excel_sentences = parse_workbook(workbook_path)
    extra_vocab, extra_sentences = build_supplemental_entries()
    bulk_vocab_raw, bulk_sentences_raw = generate_bulk_entries(deduplicate_entries(excel_vocab + extra_vocab))
    bulk_vocab = [
        make_entry(
            kind="vocab",
            source="generated-bulk",
            sheet="대량 생성 단어",
            index=index,
            thai=item["thai"],
            thai_script=item.get("thaiScript", ""),
            korean=item["korean"],
            tags=item.get("tags"),
            note=item.get("note", ""),
            extra_keywords=item.get("keywords"),
        )
        for index, item in enumerate(bulk_vocab_raw, start=1)
    ]
    bulk_sentences = [
        make_entry(
            kind="sentence",
            source="generated-bulk",
            sheet="대량 생성 문장",
            index=index,
            thai=item["thai"],
            thai_script=item.get("thaiScript", ""),
            korean=item["korean"],
            tags=item.get("tags"),
            note=item.get("note", ""),
            extra_keywords=item.get("keywords"),
        )
        for index, item in enumerate(bulk_sentences_raw, start=1)
    ]
    vocab_entries = deduplicate_entries(excel_vocab + extra_vocab + bulk_vocab)
    sentence_entries = deduplicate_entries(excel_sentences + extra_sentences + bulk_sentences)
    return {
        "appTitle": "태국어 포켓북",
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "transliterationStyle": "practical-ko",
        "note": "보강 표현의 발음 표기는 한국어 사용자 기준의 실전용 표기이며, 기존 단어를 씨앗으로 만든 대량 조합형 데이터까지 함께 검색합니다.",
        "scenarios": SCENARIOS,
        "stats": {
            "excelVocab": len(excel_vocab),
            "excelSentences": len(excel_sentences),
            "supplementalVocab": len(extra_vocab),
            "supplementalSentences": len(extra_sentences),
            "generatedBulkVocab": len(bulk_vocab),
            "generatedBulkSentences": len(bulk_sentences),
            "totalVocab": len(vocab_entries),
            "totalSentences": len(sentence_entries),
        },
        "vocab": vocab_entries,
        "sentences": sentence_entries,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Build Thai phrasebook data.js from Excel.")
    parser.add_argument(
        "--input",
        default=r"D:\Development\thai_vocab.xlsx",
        help="Path to the Excel workbook.",
    )
    parser.add_argument(
        "--output",
        default=r"D:\Development\thai_phrase_web\app\data.js",
        help="Path to the generated data.js file.",
    )
    args = parser.parse_args()

    input_path = Path(args.input)
    output_path = Path(args.output)
    data = build_data(input_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        "window.BASE_DATA = " + json.dumps(data, ensure_ascii=False, indent=2) + ";\n",
        encoding="utf-8",
    )
    print(
        json.dumps(
            {
                "input": str(input_path),
                "output": str(output_path),
                "stats": data["stats"],
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
